from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox


from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('Student_data')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Fathers Occupation"
    sheet['L1'] = "Mothers Occupation"

    file.save('Student_data')


def Exit():
    root.destroy()
###############################ShowImage###########################
def showimage():

    global filename
    global img

    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetype=(("JPG FIle","*.jpg"),
                                                                               ("PNG File","*.png"),
                                                                               ("All files","*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

#######################Regustration NO.########################
#now each time we have to enter registrartion number

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value = sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")

########################clear##########################

def Clear():

    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")


    registration_no()
    saveButton.config(state = 'normal')

    img1=PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ""

def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    print(R1)
    print(N1)
    print(C1)
    print(G1)
    print(D2)
    print(D1)
    print(Re1)
    print(S1)
    print(fathername)
    print(mothername)
    print(F1)
    print(M1)



def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"
#top frames
Label(root, text="Email: kaptur2601@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#f36464",fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

#search box to update

Search = StringVar()
Entry(root, textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
iamgeicon3 = PhotoImage(file="Images/search.png")
Srch = Button(root,text='Search', compound=LEFT, image=iamgeicon3,width=123,bg='#68ddfa', font="arial 13 bold")
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="Images/Layer 4.png")
Update_button = Button(root,image=imageicon4, bg = "#c36464")
Update_button.place(x=10,y=64)

#Registration and Date

Label(root, text="Registration No:", font="arial 13", fg=framebg,bg=background).place(x=30,y=150)
Label(root, text="Date:", font="arial 13", fg=framebg,bg=background).place(x=500,y=150)

Registration = StringVar()
Date = StringVar()


reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)


#Student details

obj=LabelFrame(root, text="Student's Details", font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:", font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:", font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:", font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:", font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:", font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:", font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name = StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)


DOB = StringVar()
DOB_entry = Entry(obj,textvariable=DOB,width=20,font="arial 10")
DOB_entry.place(x=160,y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)


Religion = StringVar()
Religion_entry = Entry(obj,textvariable=Religion,width=20,font="arial 10")
Religion_entry.place(x=630,y=100)

Skill = StringVar()
Skill_entry = Entry(obj,textvariable=Skill,width=20,font="arial 10")
Skill_entry.place(x=630,y=150)

Class = Combobox(obj, values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10", width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")




#Parents details

obj2=LabelFrame(root, text="Parent's Details", font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)


Label(obj2, text= "Fathers Name:", font="arial 13", bg=framebg, fg = framefg).place(x=30,y=50)
Label(obj2, text= "Occupation:", font="arial 13", bg=framebg, fg = framefg).place(x=30,y=100)


F_Name = StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)


Father_Occupation = StringVar()
FO_entry = Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)

Label(obj2, text= "Mothers Name:", font="arial 13", bg=framebg, fg = framefg).place(x=500,y=50)
Label(obj2, text= "Occupation:", font="arial 13", bg=framebg, fg = framefg).place(x=500,y=100)


M_Name = StringVar()
m_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=50)


Mother_Occupation = StringVar()
MO_entry = Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)


#image

f = Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img = PhotoImage(file="Images/upload photo.png")
lbl = Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button

Button(root,text="Upload",width=19,height=2,font="arial 12 bold", bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton = Button(root,text="Save",width=19,height=2,font="arial 12 bold", bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold", bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold", bg="grey", command=exit).place(x=1000,y=610)



















root.mainloop()